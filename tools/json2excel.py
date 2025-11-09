import json
from collections.abc import Generator
from typing import Any

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage

import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


class Json2ExcelTool(Tool):
    """
    Convert JSON to Excel with optional formatting support.

    Supports the [format] reserved key for row heights and column widths.
    Excel prohibits sheet names containing [ ] characters, so [format] is safe.
    """

    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage]:
        json_str = tool_parameters['json_str']

        # Parse JSON and extract sheets data and format configuration
        payload = self._load_json(json_str)
        sheets_data, format_cfg = self._extract_sheets_and_format(payload)
        defaults, sheet_formats, warnings = self._prepare_format_sections(format_cfg, sheets_data.keys())

        # Create Excel file
        excel_buffer = BytesIO()
        try:
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                for sheet_name, records in sheets_data.items():
                    if not isinstance(records, list):
                        raise Exception(f"Value for sheet '{sheet_name}' must be a list of records.")

                    # Write data to sheet
                    df = pd.DataFrame(records)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Apply formatting if configured
                    worksheet = writer.sheets[sheet_name]
                    self._apply_formatting(
                        worksheet=worksheet,
                        sheet_name=sheet_name,
                        defaults=defaults,
                        sheet_format=sheet_formats.get(sheet_name, {})
                    )
        except Exception as e:
            raise Exception(f"Error converting data to Excel: {str(e)}")

        # Create blob message with the Excel bytes
        try:
            excel_buffer.seek(0)
            excel_bytes = excel_buffer.getvalue()
            filename = tool_parameters.get('filename', 'Converted Data')
            filename = f"{filename.replace(' ', '_')}.xlsx"

            # Output warnings if any
            if warnings:
                yield self.create_text_message(f"⚠️ Warning: {warnings}")

            yield self.create_text_message(f"Excel file '{filename}' generated successfully")

            yield self.create_blob_message(
                blob=excel_bytes,
                meta={
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "filename": filename
                }
            )
        except Exception as e:
            raise Exception(f"Error creating Excel file message: {str(e)}")

    def _load_json(self, json_str: str) -> Any:
        """Parse JSON string and return the payload."""
        try:
            return json.loads(json_str)
        except json.JSONDecodeError as exc:
            raise Exception(f"Invalid JSON format: {exc}")

    def _extract_sheets_and_format(self, payload: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        """
        Extract sheets data and format configuration from payload.

        Returns:
            (sheets_data, format_cfg) where:
            - sheets_data: dict mapping sheet names to record lists
            - format_cfg: dict containing format configuration
        """
        if isinstance(payload, list):
            # Single sheet without formatting
            return {"Sheet1": payload}, {}

        if isinstance(payload, dict):
            # Extract format configuration
            # Check if [format] key exists and validate its type
            if "[format]" in payload:
                format_cfg = payload["[format]"]
                if not isinstance(format_cfg, dict):
                    raise Exception("The '[format]' section must be an object when provided.")
            else:
                format_cfg = {}

            # Extract sheets data (all keys except [format])
            sheets = {k: v for k, v in payload.items() if k != "[format]"}
            if not sheets:
                raise Exception("At least one sheet must be provided alongside the '[format]' block.")

            return sheets, format_cfg

        raise Exception("JSON must be an array (single sheet) or object (multiple sheets).")

    def _prepare_format_sections(
        self,
        format_cfg: dict[str, Any],
        sheet_names: set[str]
    ) -> tuple[dict[str, Any], dict[str, Any], str]:
        """
        Validate and prepare format configuration sections.

        Returns:
            (defaults, sheet_formats, warnings) where:
            - defaults: global default formatting
            - sheet_formats: per-sheet formatting overrides
            - warnings: warning message if any (empty string if no warnings)
        """
        if not format_cfg:
            return {}, {}, ""

        # Extract and normalize defaults and sheet_formats
        defaults = format_cfg.get("defaults")
        sheet_formats = format_cfg.get("sheets")

        # Validate types and normalize falsy values to empty dicts
        if "defaults" in format_cfg:
            if defaults is None:
                defaults = {}
            elif not isinstance(defaults, dict):
                raise Exception("The '[format].defaults' section must be an object.")
        else:
            defaults = {}

        if "sheets" in format_cfg:
            if sheet_formats is None:
                sheet_formats = {}
            elif not isinstance(sheet_formats, dict):
                raise Exception("The '[format].sheets' section must be an object.")
        else:
            sheet_formats = {}

        # Check for unknown sheet references (warning mode)
        unknown = set(sheet_formats.keys()) - set(sheet_names)
        warning_msg = ""
        if unknown:
            missing = ", ".join(sorted(unknown))
            warning_msg = f"The '[format].sheets' section references unknown sheets: {missing}. These configurations were ignored."
            # Remove unknown sheets from sheet_formats
            for unknown_sheet in unknown:
                del sheet_formats[unknown_sheet]

        # Validate each sheet format is an object
        for sheet_name, cfg in sheet_formats.items():
            if cfg is None:
                sheet_formats[sheet_name] = {}
            elif not isinstance(cfg, dict):
                raise Exception(f"The '[format].sheets.{sheet_name}' section must be an object.")

        return defaults, sheet_formats, warning_msg

    def _apply_formatting(
        self,
        worksheet,
        sheet_name: str,
        defaults: dict[str, Any],
        sheet_format: dict[str, Any]
    ) -> None:
        """
        Apply formatting to worksheet following priority rules:
        1. defaults.rowHeight/columnWidth (global defaults)
        2. sheets.<name>.rowHeight/columnWidth (sheet defaults)
        3. defaults.rowHeights/columnWidths (global specific rows/columns)
        4. sheets.<name>.rowHeights/columnWidths (sheet specific rows/columns)
        """
        max_row = max(worksheet.max_row, 1)
        max_col = max(worksheet.max_column, 1)

        # Apply vertical center alignment to all cells by default
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(vertical='center')

        # Apply uniform row heights (priority 1 & 2)
        self._apply_uniform_row_height(
            worksheet, max_row, defaults.get("rowHeight"), "defaults.rowHeight"
        )
        self._apply_uniform_row_height(
            worksheet, max_row, sheet_format.get("rowHeight"), f"sheets.{sheet_name}.rowHeight"
        )

        # Apply uniform column widths (priority 1 & 2)
        self._apply_uniform_column_width(
            worksheet, max_col, defaults.get("columnWidth"), "defaults.columnWidth"
        )
        self._apply_uniform_column_width(
            worksheet, max_col, sheet_format.get("columnWidth"), f"sheets.{sheet_name}.columnWidth"
        )

        # Apply specific row heights (priority 3 & 4)
        self._apply_row_map(
            worksheet, defaults.get("rowHeights"), "defaults.rowHeights"
        )
        self._apply_row_map(
            worksheet, sheet_format.get("rowHeights"), f"sheets.{sheet_name}.rowHeights"
        )

        # Apply specific column widths (priority 3 & 4)
        self._apply_column_map(
            worksheet, defaults.get("columnWidths"), "defaults.columnWidths"
        )
        self._apply_column_map(
            worksheet, sheet_format.get("columnWidths"), f"sheets.{sheet_name}.columnWidths"
        )

    def _apply_uniform_row_height(self, worksheet, row_count: int, value: Any, label: str) -> None:
        """Apply the same height to all rows."""
        height = self._coerce_positive_number(value, label)
        if height is None:
            return
        for row in range(1, row_count + 1):
            worksheet.row_dimensions[row].height = height

    def _apply_uniform_column_width(self, worksheet, column_count: int, value: Any, label: str) -> None:
        """Apply the same width to all columns."""
        width = self._coerce_positive_number(value, label)
        if width is None:
            return
        for col_idx in range(1, column_count + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = width

    def _apply_row_map(self, worksheet, mapping: Any, label: str) -> None:
        """Apply specific heights to individual rows."""
        if mapping is None:
            return
        if not isinstance(mapping, dict):
            raise Exception(f"The '{label}' section must be an object.")

        for raw_row, raw_height in mapping.items():
            row_index = self._parse_row_identifier(raw_row, label)
            height = self._coerce_positive_number(raw_height, f"{label}[{raw_row}]")
            worksheet.row_dimensions[row_index].height = height

    def _apply_column_map(self, worksheet, mapping: Any, label: str) -> None:
        """Apply specific widths to individual columns."""
        if mapping is None:
            return
        if not isinstance(mapping, dict):
            raise Exception(f"The '{label}' section must be an object.")

        for raw_col, raw_width in mapping.items():
            column_letter = self._parse_column_identifier(raw_col, label)
            width = self._coerce_positive_number(raw_width, f"{label}[{raw_col}]")
            worksheet.column_dimensions[column_letter].width = width

    def _parse_row_identifier(self, raw_value: Any, label: str) -> int:
        """
        Parse row identifier to 1-based integer.

        Accepts: "1", "2", 1, 2, etc.
        """
        try:
            row = int(str(raw_value))
        except (TypeError, ValueError):
            raise Exception(f"Row identifiers in '{label}' must be positive integers.")

        if row <= 0:
            raise Exception(f"Row identifiers in '{label}' must be 1-based positive integers.")

        return row

    def _parse_column_identifier(self, raw_value: Any, label: str) -> str:
        """
        Parse column identifier to Excel letter format.

        Accepts:
        - Letters: "A", "B", "AA", "AB", etc.
        - 1-based integers: 1, 2, "1", "2", etc.

        Returns: Excel column letter (e.g., "A", "B", "AA")
        """
        if isinstance(raw_value, int):
            index = raw_value
        elif isinstance(raw_value, str):
            token = raw_value.strip()
            if not token:
                raise Exception(f"Column identifiers in '{label}' cannot be empty.")

            if token.isdigit():
                # Numeric string like "1", "2"
                index = int(token)
            elif token.isalpha():
                # Letter string like "A", "B", "AA"
                return token.upper()
            else:
                raise Exception(f"Column identifiers in '{label}' must be letters or integers.")
        else:
            raise Exception(f"Column identifiers in '{label}' must be letters or integers.")

        # Convert 1-based index to Excel letter
        if index <= 0:
            raise Exception(f"Column identifiers in '{label}' must be 1-based positive integers.")

        return get_column_letter(index)

    def _coerce_positive_number(self, value: Any, label: str) -> float | None:
        """
        Coerce value to a positive number.

        Returns None if value is None, otherwise validates and returns float.
        """
        if value is None:
            return None

        try:
            number = float(value)
        except (TypeError, ValueError):
            raise Exception(f"The value for '{label}' must be a positive number.")

        if number <= 0:
            raise Exception(f"The value for '{label}' must be greater than zero.")

        return number
